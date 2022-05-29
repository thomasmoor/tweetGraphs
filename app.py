# snscrape
# https://www.youtube.com/watch?v=jtIMnmbnOFo
# Sentiment:   https://www.youtube.com/watch?v=uPKnSq6TaAk
# df Text cleansing: https://www.youtube.com/watch?v=ujId4ipkBio
import base64
import datetime
from flask import Flask,jsonify,redirect,render_template,request,Response,session,url_for
from flask_cors import CORS, cross_origin   # pip install -U flask-cors
from flask_session import Session           # pip install -U flask-session
from io import BytesIO
import json
import logging
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.writer.excel import save_virtual_workbook
import pandas as pd
import re
from scipy.special import softmax
import snscrape.modules.twitter as sntwitter
from textblob import TextBlob
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import urllib
from wordcloud import WordCloud

# pip install flask flask-cors flask-session matplotlib nltk pandas scipy snscrape textblob transformers torch wordcloud

# nltk.download('stopwords')
days=10
limit=100
stops = set(stopwords.words('english'))
sentiment_model = "cardiffnlp/twitter-roberta-base-sentiment"

#
# Logging
#
logging.basicConfig(
  filename='TweetGraphs.log',
  # encoding='utf-8', 
  format='%(asctime)s %(levelname)s:%(message)s', 
  level=logging.DEBUG
)

app = Flask(__name__)

#
# Cross Origin to be able to call the API from another machine or port
#
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

# Use Flask server session to avoid a "Confirm Form Resubmission" pop-up:
# Redirect and pass form values from post to get method
app.config['SECRET_KEY'] = "your_secret_key" 
app.config['SESSION_TYPE'] = 'filesystem' 
app.config['SESSION_PERMANENT']= False
app.config.from_object(__name__)
Session(app)

@app.route('/gettweetgraphs', methods=['POST'])
@cross_origin()
def api():
  logging.debug(f"/gettweetgraphs - got request: {request.method}")
  data = json.loads(request.data)
  
  # Get and process tweets
  tweets_process(data)
  
  if 'tweets' in session:
    res={
      'tweets': session['tweets'],
      'word_graph': session['word_graph'],
      'sentiment_graph': session['sentiment_graph'],
    }
  else:
    res={
      'tweets': [],
      'word_graph': '',
      'sentiment_graph': '',
    }
  return jsonify(res)
# api

@app.route('/', methods=['GET','POST'])
@cross_origin()
def slash():
  # print(f"slash - got request: {request.method}")
  logging.debug(f"slash - got request: {request.method}")
  logging.debug(f"r.f: {request.form}")
  # print(request.form)

  # The 'extract' button was pressed
  if 'extract' in request.form:
    logging.debug("extract")    
    # data = json.loads(request.data)
    # print("data:")
    # print(data)
    # search=data['keyword']
    # user=data['user']
    data=request.form
    tweets_process(data)
  # extract
  
  # Download Option
  elif 'download' in request.form and 'tweets' in session:
  
    # Create a workbook
    wb = Workbook()
    
    # Assign the sheets
    ws = wb.active
    ws.title = "Tweets"
    # print(ws.title)
    
    # Get the data
    j=session['tweets']
    if j:
      tweets=session['tweets']
    else:
      tweets=[]
    
    # Set the sheets
    set_sheets(tweets,ws)
    
    return Response(
      save_virtual_workbook(wb),
      headers={
        'Content-Disposition': 'attachment; filename=Tweets.xlsx',
        'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
      }
    ) 
  # download
    
  # Redirect
  if request.method=='POST':
    logging.debug("TweetGraphs - branch: redirect")
    return redirect(url_for('slash'))

  # Render
  else:
    logging.debug("TweetGraphs - branch: render index.html")
    if 'tweets' in session:
        tweets=session['tweets']
    else:
      tweets=[]
    logging.debug("TweetGraphs - Rendered tweets:")
    # logging.debug(df)

    if 'word_graph' in session:
        word_graph=session['word_graph']
    else:
      word_graph=None

    if 'sentiment_graph' in session:
        sentiment_graph=session['sentiment_graph']
    else:
      sentiment_graph=None

    return render_template("index.html",tweets=tweets,sentiment_graph=sentiment_graph,word_graph=word_graph)
# slash

def clean_text(t):
  t=t.lower()
  t=re.sub(r'@[A-Za-z0-9]+','',t)
  t=re.sub(r'#','',t)
  t=re.sub(r'RT[\s]+','',t)  # Retweet
  # t=re.sub(r'https?:\/\/\S+','http',t)
  t=re.sub(r'https?:\/\/\S+','',t)
  return t
# clean_text

def get_graph():
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0) # Go to the beginning of the buffer
    image_png=buffer.getvalue()
    graph=base64.b64encode(image_png)
    graph=graph.decode('utf-8')
    buffer.close()
    return graph
# get_graph
	
def get_polarity(text):
  return TextBlob(text).sentiment.polarity
# get_polarity

def get_subjectivity(text):
  return TextBlob(text).sentiment.subjectivity
# get_subjectivity

def scatter_polarity_subjectivity(df):
  plt.figure(figsize=(8,6))
  for i in range(0,df.shape[0]):
    plt.scatter(df['Sentiment'][i],df['Subjectivity'][i],color='Blue')
  plt.title("Sentiment / Subjectivity")
  plt.xlabel('Sentiment')
  plt.ylabel('Subjectivity')
  # plt.show()
# scatter_polarity_subjectivity

def sentiment(tweet):
  content=tweet[2]
  print(f"Sentiment of: {tweet[0]} - {tweet[1]}: {tweet[2]}")
  
  # preprocess tweet
  tweet_words = []
  for word in content.split(' '):
    if word.startswith('@') and len(word) > 1:
        word = '@user'
    elif word.startswith('http'):
        word = "http"
    tweet_words.append(word)

  tweet_proc = " ".join(tweet_words)

  labels = ['Negative', 'Neutral', 'Positive']

  # sentiment analysis
  encoded_tweet = tokenizer(tweet_proc, return_tensors='pt')
  # print("Encode Tweet:")
  # print(encoded_tweet)
  # output = model(encoded_tweet['input_ids'], encoded_tweet['attention_mask'])
  output = model(**encoded_tweet)
  # print("Model Output:")
  # print(output)

  scores = output[0][0].detach().numpy()
  # print("Scores:")
  # print(scores)
  scores = softmax(scores)
  # print("Softmax scores:")
  # print(scores)
  
  for i in range(len(scores)):  
    l = labels[i]
    s = scores[i]
    print(l,s)
# sentiment

def set_sheets(data,ws):

    c=1
    r=1
      
    ws[get_column_letter(c)+str(r)]='Date'
    ws[get_column_letter(c+1)+str(r)]='User'
    ws[get_column_letter(c+2)+str(r)]='Tweet'
    ws[get_column_letter(c+3)+str(r)]='Like'
    ws[get_column_letter(c+4)+str(r)]='Reply'
    ws[get_column_letter(c+5)+str(r)]='Retweet'
    ws[get_column_letter(c+6)+str(r)]='Sentiment'
    ws[get_column_letter(c+7)+str(r)]='Subjectivity'

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    # Data
    logging.debug("set_sheet - data")
    logging.debug(data)
    for row in data:
      logging.debug("set_sheet - row")
      logging.debug(row)
      r+=1
      ws[get_column_letter(c)+str(r)]=f"{row['Date']}"
      ws[get_column_letter(c+1)+str(r)]=row['User']
      ws[get_column_letter(c+2)+str(r)]=row['Tweet']
      ws[get_column_letter(c+3)+str(r)]=row['LikeCount']
      ws[get_column_letter(c+4)+str(r)]=row['ReplyCount']
      ws[get_column_letter(c+5)+str(r)]=row['RetweetCount']
      ws[get_column_letter(c+6)+str(r)]=row['Sentiment']
      ws[get_column_letter(c+7)+str(r)]=row['Subjectivity']
        
# set_sheets

def tweets_process(d):
  print(f"tweets - {d}")
  
  # Get the search parameters
  query=d['query']
  lang=d['lang']
  if 'min_faves' in d:
    min_faves=d['min_faves']
  else:
    min_faves=None
  if 'min_replies' in d:
    min_replies=d['min_replies']
  else:
    min_replies=None
  if 'min_retweets' in d:
    min_retweets=d['min_retweets']
  else:
    min_retweets=None
  search=d['keyword']
  user=d['user']
  # logging.debug(f"tweets - user: {user} search:{search} query:{query}")
  
  # Create the query
  date_from=datetime.datetime.now() - datetime.timedelta(days=days)
  date_until=None
  q=tweets_query(date_from,date_until,lang,min_faves,min_replies,min_retweets,query,search,user)
  print("q: "+q)
  
  # Download the tweets
  df=tweets_download(q)
  logging.debug("tweets - df:")
  logging.debug(f"df: {df.shape}")
  
  # Cleanse the data
  df['Tweet']=df['Tweet'].apply(clean_text)
 
  # Calculate the Sentiment
  df['Sentiment']=df['Tweet'].apply(get_polarity)
  # Calculate the Subjectivity
  df['Subjectivity']=df['Tweet'].apply(get_subjectivity)

  # print(df)
  # Save the data into Session
  if not df.empty:
    # Tweets
    session['tweets'] = df.to_dict('records')  
    # Graph: word cloud
    word_cloud(df,700,1000,5,10,"")
    word_graph=get_graph()
    session['word_graph']=word_graph
    # Graph: sentiment scatter    
    scatter_polarity_subjectivity(df)
    sentiment_graph=get_graph()
    session['sentiment_graph']=sentiment_graph
  else:
    session['tweets'] =[]
    session['word_graph']=''
    session['sentiment_graph']=''
    
# tweets

# Tweets - Download
def tweets_download(q):

  # Get tweets
  logging.debug(f"tweets_download: {q}")
  print(f"tweets_download: {q}")
  tweets=[]
  for t in sntwitter.TwitterSearchScraper(q).get_items():
    # print(f"tweets_download {t.date} - {t.user.username}: {t.content}")
    # print(vars(t))
    if len(tweets)==limit:
      break
    tweets.append([t.date,t.user.username,t.content,t.url,t.likeCount,t.replyCount,t.retweetCount])
    
  logging.debug(f"tweets_download - size of tweets: {len(tweets)}")
  print(f"tweets_download - size of tweets: {len(tweets)}")
  
  # Transform into a Pandas dataframe  
  df=pd.DataFrame(tweets,columns=['Date','User','Tweet','URL','LikeCount','ReplyCount','RetweetCount'])
  # logging.debug(f"tweets_download - df:")
  # logging.debug(df)
  print(df.shape)
  
  return df    
# tweets_download

def tweets_query(date_since,date_until,lang,min_faves,min_replies,min_retweets,query,search,user_from):
  # falcon%20(from%3Aelonmusk)%20since%3a2022-05-25%20

  search=search.lower()
  
  # Manual Query
  if query and query!="":
    query=query.lower()
    # logging.debug(f"Manual Query - query 1={query}")
    query=re.sub('(since:\d*-\d*-\d*)','since:'+date_since.strftime('%Y-%m-%d'),query)
    # logging.debug(f"Manual Query - query 2={query}")
    return query
  
  if not lang or lang=="": lang='en'
  q="lang:"+lang
  sep=" "
  if search and search!="":
    # search=urllib.parse.quote(search)
    q+=f" {search}"

  if user_from and user_from!="":
    q+=f" (from:{user_from})"

  if date_since and date_since!="":
    q+=f" since:{date_since.strftime('%Y-%m-%d')}"

  if date_until and date_until!="":
    q+=f" until:{date_until.strftime('%Y-%m-%d')}"
  
  if min_replies and min_replies!="":
    q+=f" min_replies:{min_replies}"
  
  if min_faves and min_faves!="":
    q+=f" min_faves:{min_faves}"
  
  if min_retweets and min_retweets!="":
    q+=f" min_retweets:{min_retweets}"
  
  logging.debug(f"q:{q}")
  return q
# tweets_query

def word_cloud(df,cloud_height,cloud_width,fig_height,fig_width,title):
  allWords=' '.join( [t for t in df['Tweet']])
  # print(allWords)
  # filtered=''
  # for w in allWords:
  #  # print(w)
  #  if w not in stops:
  #    filtered+=w+' '
  # print(f"allwords: {len(allWords)}")
  # print(f"stops: {len(stops)}")
  # print(f"filtered: {len(filtered)}")
  cloud=WordCloud(width=cloud_width,height=cloud_height,random_state=21,max_font_size=119,stopwords=stops).generate(allWords)
  # cloud=WordCloud(random_state=21,max_font_size=119,stopwords=stops).generate(allWords)
  plt.imshow(cloud,interpolation="bilinear")
  # plt.figure(figsize=(fig_width,fig_height))
  plt.title(title)
  plt.axis('off')
  plt.margins(x=0, y=0)
  # plt.show()
# word_cloud

if __name__ == "__main__":

  # Start the web server
  app.run(debug=True,host='0.0.0.0',port=5004)